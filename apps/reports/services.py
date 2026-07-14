from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.pricing.models import CollectionRun, ComparisonGroup, ComparisonItem

STATUS_FILL = {
    ComparisonItem.ColorStatus.GREEN: "C6EFCE",
    ComparisonItem.ColorStatus.YELLOW: "FFEB9C",
    ComparisonItem.ColorStatus.ORANGE: "F4B183",
    ComparisonItem.ColorStatus.RED: "FFC7CE",
}


class ExcelReportService:
    @staticmethod
    def _set_widths(sheet, widths: dict[int, float]) -> None:
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width

    @classmethod
    def export_run(cls, run: CollectionRun, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Сравнение"
        details = workbook.create_sheet("Детали")

        groups = list(
            run.comparison_groups.select_related(
                "hotel", "room_category", "meal_plan", "best_offer__source"
            ).prefetch_related("items__offer__source")
        )
        source_names = sorted(
            {item.offer.source.name for group in groups for item in group.items.all()}
        )

        summary["A1"] = "Сравнение цен туроператоров"
        summary["A1"].font = Font(size=16, bold=True)
        summary.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=10 + len(source_names)
        )
        summary["A2"] = f"Сценарий: {run.scenario.name}"
        summary["A3"] = (
            f"Дата заезда: {run.scenario.check_in:%d.%m.%Y}; ночей: {run.scenario.nights}; взрослых: {run.scenario.adults}"
        )

        base_headers = [
            "Отель",
            "Категория номера",
            "Питание",
            "Дата заезда",
            "Ночей",
            "Лучшая цена",
            "Валюта",
            "Лучший источник",
        ]
        end_headers = ["Макс. отклонение", "Комментарий"]
        headers = base_headers + source_names + end_headers
        header_row = 5
        for column, value in enumerate(headers, 1):
            cell = summary.cell(header_row, column, value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        summary.freeze_panes = "A6"
        summary.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        )

        for row_number, group in enumerate(groups, header_row + 1):
            items = {item.offer.source.name: item for item in group.items.all()}
            max_percent = max(
                (item.percent_difference for item in items.values()), default=0
            )
            values = [
                group.hotel.canonical_name,
                group.room_category.canonical_name,
                group.meal_plan.code,
                group.check_in,
                group.nights,
                group.best_price,
                group.currency,
                group.best_offer.source.name,
            ]
            for column, value in enumerate(values, 1):
                summary.cell(row_number, column, value)

            source_start = len(base_headers) + 1
            for offset, source_name in enumerate(source_names):
                cell = summary.cell(row_number, source_start + offset)
                item = items.get(source_name)
                if item:
                    cell.value = item.offer.price
                    cell.fill = PatternFill(
                        "solid", fgColor=STATUS_FILL[item.color_status]
                    )
                    cell.number_format = "#,##0.00"
                    cell.comment = None
                else:
                    cell.value = "Нет цены"
                    cell.fill = PatternFill("solid", fgColor="D9D9D9")

            summary.cell(
                row_number, source_start + len(source_names), max_percent / 100
            )
            summary.cell(
                row_number, source_start + len(source_names)
            ).number_format = "0.00%"
            summary.cell(
                row_number,
                source_start + len(source_names) + 1,
                "Сравниваются только полностью совпадающие параметры",
            )

        cls._set_widths(
            summary,
            {
                1: 34,
                2: 28,
                3: 12,
                4: 14,
                5: 9,
                6: 15,
                7: 10,
                8: 20,
                **{9 + i: 18 for i in range(len(source_names))},
                9 + len(source_names): 17,
                10 + len(source_names): 36,
            },
        )
        for row in summary.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(header_row + 1, summary.max_row + 1):
            summary.cell(row, 4).number_format = "DD.MM.YYYY"
            summary.cell(row, 6).number_format = "#,##0.00"

        detail_headers = [
            "Источник",
            "Отель",
            "Категория",
            "Питание",
            "Заезд",
            "Ночей",
            "Цена",
            "Валюта",
            "Разница",
            "Разница, %",
            "Цвет",
            "Налоги включены",
            "Трансфер включён",
            "Ссылка",
        ]
        details.append(detail_headers)
        for cell in details[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for group in groups:
            for item in group.items.all():
                details.append(
                    [
                        item.offer.source.name,
                        group.hotel.canonical_name,
                        group.room_category.canonical_name,
                        group.meal_plan.code,
                        group.check_in,
                        group.nights,
                        item.offer.price,
                        group.currency,
                        item.absolute_difference,
                        item.percent_difference / 100,
                        item.get_color_status_display(),
                        "Да" if group.taxes_included else "Нет",
                        "Да" if group.transfer_included else "Нет",
                        item.offer.offer_url,
                    ]
                )
                current_row = details.max_row
                for column in (7, 9):
                    details.cell(current_row, column).number_format = "#,##0.00"
                details.cell(current_row, 10).number_format = "0.00%"
                details.cell(current_row, 11).fill = PatternFill(
                    "solid", fgColor=STATUS_FILL[item.color_status]
                )
                details.cell(current_row, 5).number_format = "DD.MM.YYYY"

        details.freeze_panes = "A2"
        details.auto_filter.ref = f"A1:N{details.max_row}"
        cls._set_widths(
            details,
            {
                1: 20,
                2: 34,
                3: 28,
                4: 12,
                5: 14,
                6: 9,
                7: 15,
                8: 10,
                9: 15,
                10: 14,
                11: 18,
                12: 17,
                13: 18,
                14: 40,
            },
        )
        for row in details.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        workbook.save(output_path)
        return output_path


class PdfReportService:
    @staticmethod
    def _register_font() -> str:
        candidates = [
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
        ]
        for path in candidates:
            if path.exists():
                pdfmetrics.registerFont(TTFont("ReportFont", str(path)))
                return "ReportFont"
        return "Helvetica"

    @classmethod
    def export_run(cls, run: CollectionRun, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        font_name = cls._register_font()

        document = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title="Сравнение цен туроператоров",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "RussianTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=16,
            leading=19,
            alignment=TA_CENTER,
        )
        normal_style = ParagraphStyle(
            "RussianNormal",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=8,
            leading=10,
        )

        story = [
            Paragraph("Сравнение цен туроператоров", title_style),
            Paragraph(
                f"Сценарий: {run.scenario.name}. Заезд: {run.scenario.check_in:%d.%m.%Y}; "
                f"ночей: {run.scenario.nights}; взрослых: {run.scenario.adults}.",
                normal_style,
            ),
            Spacer(1, 5 * mm),
        ]

        data = [
            [
                Paragraph("Отель", normal_style),
                Paragraph("Категория", normal_style),
                Paragraph("Питание", normal_style),
                Paragraph("Лучшая цена", normal_style),
                Paragraph("Лучший источник", normal_style),
                Paragraph("Макс. отклонение", normal_style),
            ]
        ]

        groups: Iterable[ComparisonGroup] = run.comparison_groups.select_related(
            "hotel", "room_category", "meal_plan", "best_offer__source"
        ).prefetch_related("items")
        for group in groups:
            max_percent = max(
                (item.percent_difference for item in group.items.all()), default=0
            )
            data.append(
                [
                    Paragraph(group.hotel.canonical_name, normal_style),
                    Paragraph(group.room_category.canonical_name, normal_style),
                    Paragraph(group.meal_plan.code, normal_style),
                    Paragraph(
                        f"{group.best_price:,.2f} {group.currency}", normal_style
                    ),
                    Paragraph(group.best_offer.source.name, normal_style),
                    Paragraph(f"{max_percent:.2f}%", normal_style),
                ]
            )

        table = Table(
            data,
            colWidths=[70 * mm, 50 * mm, 20 * mm, 35 * mm, 45 * mm, 35 * mm],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 5 * mm))
        story.append(
            Paragraph(
                "Важно: в одной строке сравниваются только совпадающие отель, категория номера, питание, "
                "дата, ночи, состав туристов, налоги, трансфер и валюта.",
                normal_style,
            )
        )
        document.build(story)
        return output_path
