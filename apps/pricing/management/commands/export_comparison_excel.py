from collections import defaultdict
from pathlib import Path

from django.conf import settings
from django.core.management.base import (
    BaseCommand,
    CommandError,
)
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from apps.pricing.models import (
    CollectionRun,
    ComparisonGroup,
    ComparisonItem,
)

SOURCE_COLUMNS = {
    "maldiviana": 12,
    "resort_holiday": 13,
}


COLOR_FILLS = {
    "green": PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    ),
    "yellow": PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    ),
    "orange": PatternFill(
        fill_type="solid",
        fgColor="F4B183",
    ),
    "red": PatternFill(
        fill_type="solid",
        fgColor="F4CCCC",
    ),
}


COLOR_FONTS = {
    "green": Font(color="006100"),
    "yellow": Font(color="9C6500"),
    "orange": Font(color="9C5700"),
    "red": Font(color="9C0006"),
}


STATUS_NAMES = {
    "green": "Одинаковая или лучшая цена",
    "yellow": "Разница до 5%",
    "orange": "Разница от 5% до 10%",
    "red": "Разница более 10%",
}


def get_rate_type(items):
    text_parts = []

    for item in items:
        offer = item.offer

        if offer.source_room_id:
            text_parts.append(offer.source_room.source_name)

        if isinstance(offer.raw_data, dict):
            text_parts.extend(
                [
                    str(
                        offer.raw_data.get(
                            "room_raw",
                            "",
                        )
                    ),
                    str(
                        offer.raw_data.get(
                            "offer_code",
                            "",
                        )
                    ),
                ]
            )

    text = " ".join(text_parts).lower()

    if "anniversary" in text:
        return "Годовщина свадьбы"

    if "honeymoon" in text:
        return "Для молодожёнов"

    return "Обычный"


def set_column_widths(worksheet, widths):
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


class Command(BaseCommand):
    help = "Создаёт Excel-отчёт по сохранённому сравнению цен"

    def add_arguments(self, parser):
        parser.add_argument(
            "--run-id",
            type=int,
            help=("ID запуска сравнения. По умолчанию используется последний."),
        )

        parser.add_argument(
            "--output",
            help=("Путь к итоговому XLSX-файлу"),
        )

    def handle(self, *args, **options):
        run_id = options["run_id"]

        if run_id is None:
            run_id = (
                ComparisonGroup.objects.order_by("-run_id")
                .values_list(
                    "run_id",
                    flat=True,
                )
                .first()
            )

        if run_id is None:
            raise CommandError("В базе пока нет сохранённых сравнений.")

        try:
            run = CollectionRun.objects.select_related("scenario").get(id=run_id)
        except CollectionRun.DoesNotExist as error:
            raise CommandError(f"Запуск ID {run_id} не найден.") from error

        groups = list(
            ComparisonGroup.objects.filter(run_id=run_id)
            .select_related(
                "hotel",
                "room_category",
                "meal_plan",
                "best_offer__source",
            )
            .order_by(
                "hotel__canonical_name",
                "room_category__canonical_name",
                "meal_plan__code",
                "id",
            )
        )

        if not groups:
            raise CommandError(f"В запуске ID {run_id} нет групп сравнения.")

        group_ids = [group.id for group in groups]

        items = list(
            ComparisonItem.objects.filter(group_id__in=group_ids)
            .select_related(
                "group",
                "offer__source",
                "offer__source_room",
            )
            .order_by(
                "group_id",
                "offer__price",
            )
        )

        items_by_group = defaultdict(list)

        for item in items:
            items_by_group[item.group_id].append(item)

        output_option = options["output"]

        if output_option:
            output_path = Path(output_option)

            if not output_path.is_absolute():
                output_path = Path(settings.BASE_DIR) / output_path
        else:
            output_path = (
                Path(settings.BASE_DIR)
                / "reports_output"
                / (f"comparison_run_{run_id}.xlsx")
            )

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Сводка"

        detail_sheet = workbook.create_sheet("Сравнение")

        dark_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        accent_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        white_bold_font = Font(
            color="FFFFFF",
            bold=True,
        )

        title_font = Font(
            bold=True,
            size=16,
            color="FFFFFF",
        )

        thin_side = Side(
            style="thin",
            color="B7B7B7",
        )

        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        # Лист «Сравнение».
        detail_sheet.merge_cells("A1:T1")

        detail_sheet["A1"] = "СРАВНЕНИЕ ЦЕН НА ОТЕЛИ"

        detail_sheet["A1"].fill = dark_fill
        detail_sheet["A1"].font = title_font
        detail_sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        detail_sheet.row_dimensions[1].height = 30

        detail_sheet["A2"] = "ID запуска"
        detail_sheet["B2"] = run.id

        detail_sheet["D2"] = "Сценарий"
        detail_sheet["E2"] = run.scenario.name if run.scenario_id else ""

        detail_sheet["A3"] = "Дата заезда"
        detail_sheet["B3"] = (
            run.scenario.check_in if run.scenario_id else groups[0].check_in
        )

        detail_sheet["B3"].number_format = "dd.mm.yyyy"

        detail_sheet["D3"] = "Количество групп"
        detail_sheet["E3"] = len(groups)

        detail_sheet["A4"] = "Валюта"
        detail_sheet["B4"] = groups[0].currency

        detail_sheet["D4"] = "Статус запуска"
        detail_sheet["E4"] = str(run.status)

        headers = [
            "№",
            "Отель",
            "Категория номера",
            "Питание",
            "Тариф",
            "Дата заезда",
            "Ночей",
            "Взрослых",
            "Возраст детей",
            "Трансфер",
            "Налоги",
            "Мальдивиана",
            "Resort Holiday",
            "Лучшая цена",
            "Валюта",
            "Лучший источник",
            "Экономия",
            "Разница, %",
            "Комментарий",
            "Ссылка",
        ]

        header_row = 6

        for column_number, header in enumerate(
            headers,
            start=1,
        ):
            cell = detail_sheet.cell(
                row=header_row,
                column=column_number,
                value=header,
            )

            cell.fill = dark_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

        detail_sheet.row_dimensions[header_row].height = 38.1

        summary_rows = []

        wins = defaultdict(int)
        equal_prices = 0
        maximum_saving = 0
        maximum_percent = 0

        for number, group in enumerate(
            groups,
            start=1,
        ):
            row = header_row + number

            # Увеличиваем высоту строки для многострочных значений.
            detail_sheet.row_dimensions[row].height = 30

            group_items = items_by_group.get(
                group.id,
                [],
            )

            if not group_items:
                continue

            prices_by_source = {
                item.offer.source.code: (item.offer.price) for item in group_items
            }

            prices = [item.offer.price for item in group_items]

            minimum_price = min(prices)
            maximum_price = max(prices)

            saving = maximum_price - minimum_price

            percent_difference = max(
                (item.percent_difference for item in group_items),
                default=0,
            )

            if all(price == minimum_price for price in prices):
                winner = "Одинаковая цена"
                equal_prices += 1
                color_status = "green"
                comment = "Стоимость предложений одинаковая"
            else:
                winner = group.best_offer.source.name

                wins[winner] += 1

                status_item = max(
                    group_items,
                    key=lambda item: item.percent_difference,
                )

                color_status = status_item.color_status

                comment = (
                    f"{winner} дешевле на "
                    f"{saving:.2f} "
                    f"{group.currency} "
                    f"({percent_difference:.2f}%)"
                )

            maximum_saving = max(
                maximum_saving,
                saving,
            )

            maximum_percent = max(
                maximum_percent,
                percent_difference,
            )

            children = (
                ", ".join(str(age) for age in (group.children_ages or [])) or "Нет"
            )

            meal_name = getattr(
                group.meal_plan,
                "name",
                "",
            )

            meal_value = f"{group.meal_plan.code}" + (
                f" — {meal_name}" if meal_name else ""
            )

            row_values = [
                number,
                group.hotel.canonical_name,
                group.room_category.canonical_name,
                meal_value,
                get_rate_type(group_items),
                group.check_in,
                group.nights,
                group.adults,
                children,
                ("Включён" if group.transfer_included else "Не включён"),
                ("Включены" if group.taxes_included else "Не включены"),
                prices_by_source.get("maldiviana"),
                prices_by_source.get("resort_holiday"),
                group.best_price,
                group.currency,
                winner,
                saving,
                percent_difference,
                comment,
                "Открыть",
            ]

            for column_number, value in enumerate(
                row_values,
                start=1,
            ):
                cell = detail_sheet.cell(
                    row=row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            detail_sheet.cell(
                row=row,
                column=6,
            ).number_format = "dd.mm.yyyy"

            for column_number in (
                12,
                13,
                14,
                17,
            ):
                detail_sheet.cell(
                    row=row,
                    column=column_number,
                ).number_format = "#,##0.00"

            detail_sheet.cell(
                row=row,
                column=18,
            ).number_format = "0.00"

            status_fill = COLOR_FILLS.get(color_status)

            status_font = COLOR_FONTS.get(color_status)

            for column_number in (
                14,
                16,
                17,
                18,
                19,
            ):
                cell = detail_sheet.cell(
                    row=row,
                    column=column_number,
                )

                if status_fill:
                    cell.fill = status_fill

                if status_font:
                    cell.font = Font(
                        color=status_font.color,
                        bold=(column_number in (14, 16)),
                    )

            for item in group_items:
                price_column = SOURCE_COLUMNS.get(item.offer.source.code)

                if price_column is None:
                    continue

                price_cell = detail_sheet.cell(
                    row=row,
                    column=price_column,
                )

                if item.offer_id == group.best_offer_id:
                    price_cell.fill = COLOR_FILLS["green"]
                    price_cell.font = Font(
                        color="006100",
                        bold=True,
                    )
                else:
                    price_cell.fill = COLOR_FILLS.get(
                        item.color_status,
                        accent_fill,
                    )

            best_url = group.best_offer.offer_url or ""

            link_cell = detail_sheet.cell(
                row=row,
                column=20,
            )

            if best_url:
                link_cell.hyperlink = best_url
                link_cell.style = "Hyperlink"
            else:
                link_cell.value = "Нет ссылки"

            summary_rows.append(
                {
                    "hotel": (group.hotel.canonical_name),
                    "winner": winner,
                    "best_price": (group.best_price),
                    "currency": group.currency,
                    "saving": saving,
                    "percent": (percent_difference),
                    "status": (
                        STATUS_NAMES.get(
                            color_status,
                            color_status,
                        )
                    ),
                    "color": color_status,
                }
            )

        last_detail_row = header_row + len(groups)

        detail_table = Table(
            displayName=(f"ComparisonRun{run.id}"),
            ref=(f"A{header_row}:T{last_detail_row}"),
        )

        detail_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        detail_sheet.add_table(detail_table)

        detail_sheet.freeze_panes = "A7"

        # Фиксируем заголовок и подготавливаем лист к печати.
        # Сетка отключена, чтобы отчёт выглядел одинаково в Excel и PDF.

        detail_sheet.sheet_view.showGridLines = False

        detail_sheet.page_setup.orientation = "landscape"

        detail_sheet.page_setup.fitToWidth = 1
        detail_sheet.page_setup.fitToHeight = 0

        set_column_widths(
            detail_sheet,
            {
                "A": 6,
                "B": 40.57,
                "C": 36,
                "D": 17,
                "E": 22,
                "F": 14,
                "G": 9,
                "H": 13.14,
                "I": 15,
                "J": 14,
                "K": 14,
                "L": 17,
                "M": 18,
                "N": 16,
                "O": 10,
                "P": 22,
                "Q": 14,
                "R": 14,
                "S": 42,
                "T": 14,
            },
        )

        # Лист «Сводка».
        summary_sheet.merge_cells("A1:G1")

        summary_sheet["A1"] = "СВОДКА ПО СРАВНЕНИЮ ЦЕН"

        summary_sheet["A1"].fill = dark_fill
        summary_sheet["A1"].font = title_font
        summary_sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        summary_sheet.row_dimensions[1].height = 30

        summary_sheet["A3"] = "Количество сравнений"
        summary_sheet["B3"] = len(groups)

        summary_sheet["A4"] = "Побед Мальдивианы"
        summary_sheet["B4"] = wins.get(
            "Мальдивиана",
            0,
        )

        summary_sheet["A5"] = "Побед Resort Holiday"
        summary_sheet["B5"] = wins.get(
            "Resort Holiday",
            0,
        )

        summary_sheet["D3"] = "Одинаковых цен"
        summary_sheet["E3"] = equal_prices

        summary_sheet["D4"] = "Максимальная экономия"
        summary_sheet["E4"] = maximum_saving
        summary_sheet["F4"] = groups[0].currency

        summary_sheet["D5"] = "Максимальная разница"
        summary_sheet["E5"] = maximum_percent
        summary_sheet["F5"] = "%"

        for row in range(3, 6):
            for column in (
                1,
                4,
            ):
                summary_sheet.cell(
                    row=row,
                    column=column,
                ).fill = accent_fill

                summary_sheet.cell(
                    row=row,
                    column=column,
                ).font = Font(bold=True)

        summary_headers = [
            "№",
            "Отель",
            "Лучший источник",
            "Лучшая цена",
            "Валюта",
            "Экономия",
            "Разница, %",
        ]

        summary_header_row = 8

        for column_number, header in enumerate(
            summary_headers,
            start=1,
        ):
            cell = summary_sheet.cell(
                row=summary_header_row,
                column=column_number,
                value=header,
            )

            cell.fill = dark_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

        for number, summary_data in enumerate(
            summary_rows,
            start=1,
        ):
            row = summary_header_row + number

            # Увеличиваем высоту строки для многострочных значений.
            summary_sheet.row_dimensions[row].height = 30

            values = [
                number,
                summary_data["hotel"],
                summary_data["winner"],
                summary_data["best_price"],
                summary_data["currency"],
                summary_data["saving"],
                summary_data["percent"],
            ]

            for column_number, value in enumerate(
                values,
                start=1,
            ):
                cell = summary_sheet.cell(
                    row=row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            summary_sheet.cell(
                row=row,
                column=4,
            ).number_format = "#,##0.00"

            summary_sheet.cell(
                row=row,
                column=6,
            ).number_format = "#,##0.00"

            summary_sheet.cell(
                row=row,
                column=7,
            ).number_format = "0.00"

            color_status = summary_data["color"]

            for column_number in (
                3,
                4,
                6,
                7,
            ):
                cell = summary_sheet.cell(
                    row=row,
                    column=column_number,
                )

                cell.fill = COLOR_FILLS.get(
                    color_status,
                    accent_fill,
                )

                cell.font = Font(
                    color=(
                        COLOR_FONTS.get(
                            color_status,
                            Font(color="000000"),
                        ).color
                    ),
                    bold=(column_number in (3, 4)),
                )

        last_summary_row = summary_header_row + len(summary_rows)

        summary_table = Table(
            displayName=(f"SummaryRun{run.id}"),
            ref=(f"A{summary_header_row}:G{last_summary_row}"),
        )

        summary_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        summary_sheet.add_table(summary_table)

        summary_sheet.freeze_panes = "A9"
        summary_sheet.sheet_view.showGridLines = False

        summary_sheet.page_setup.orientation = "landscape"

        summary_sheet.page_setup.fitToWidth = 1
        summary_sheet.page_setup.fitToHeight = 0

        set_column_widths(
            summary_sheet,
            {
                "A": 43,
                "B": 51.71,
                "C": 24,
                "D": 37.29,
                "E": 12,
                "F": 16,
                "G": 16,
            },
        )

        workbook.save(output_path)

        self.stdout.write(self.style.SUCCESS("Excel-отчёт успешно создан."))

        self.stdout.write(f"Запуск сравнения: {run.id}")

        self.stdout.write(f"Групп: {len(groups)}")

        self.stdout.write(f"Файл: {output_path}")
